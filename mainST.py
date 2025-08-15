from analysis import run_analysis
from highlight import highlight_side_lane
from delivery_helpers import generate_times
from summary import summary_delivery

import streamlit as st
import pandas as pd
import math
from datetime import datetime
from openpyxl import load_workbook
import numpy as np
from io import BytesIO
from openpyxl.styles import PatternFill



col1, col2 = st.columns([1, 6])  # adjust ratio

with col1:
    st.image("logo.png", width=100)

with col2:
    st.title("Inbound Delivery Optimization Tool")
    st.write("This tool is for optimizing move orders using BOM data and according to trailer contraints across multiple drive units")

uploaded_file = st.file_uploader("Upload BOM Excel File", type=["xlsx", "xlsm"])
if uploaded_file is None:
    st.stop()
valid_units = ['Proteus', 'Hercules', 'Megasus']
units_to_plan = st.multiselect("Select Drive Units to plan", valid_units, default=valid_units)

st.write("Max Trailer Capacity is 48 pallets with Megasus conveyor units not being able to be stacked.")


st.image("Layout.png", caption="Labeled Layout of CMA Area to go along with Excel Sheet", use_container_width=True)
if st.button("Run Analysis & Balance Deliveries"):

    combined_filename = "Delivery_Inventory_Plan_All.xlsx"
    unit_lanes = {}

    # Read BOM initial sheet
    df_bom_init = pd.read_excel(uploaded_file, sheet_name=f"Inbound-{valid_units[0]}", skiprows=15, header=None)
    df_bom_init.columns = ['Part Number', 'Description', 'Quantity / Unit', 'Needed per day',
        'Quantity Needed for Shift 1', 'Quantity Needed for Shift 2',
        'Pallets Utilized for Shift 1', 'Pallets Utilized for Shift 2',
        'Consumption Rate Units/ Hour Shift 1', 'Consumption Rate / Hour Shift 2',
        'Standard Pack Size', 'Package Type', 'Maximum Storage on Lineside',
        'Minimum Storage on Lineside', 'On-hand qty',  'In-Transit QTY', 'On-hand on dock', 'On-hand QTY at Lineside', 'Total Move Order - Prev Day', 'Total QTY Needed Previous Day']

    inventory_on_hand = {
        part: qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['On-hand qty'])
    }
    dock_on_hand = {
        part: qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['On-hand on dock'])
    }
    move_order_prev_Day = {
        part:qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['Total Move Order - Prev Day'])
    }
    # Determine max cadences
    cadence_shift_1_list, cadence_shift_2_list = [], []
    for unit in units_to_plan:
        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb[f'Inbound-{unit}']
        c1 = int(ws['B5'].value or 0)
        c2 = int(ws['B13'].value or 0)
        cadence_shift_1_list.append(c1)
        cadence_shift_2_list.append(c2)

    max_cadence_1 = max(cadence_shift_1_list)
    max_cadence_2 = max(cadence_shift_2_list)

    start_shift_1 = datetime.strptime("6:15", "%H:%M")
    end_shift_1 = datetime.strptime("15:00", "%H:%M")
    start_shift_2 = datetime.strptime("15:00", "%H:%M")
    end_shift_2 = datetime.strptime("23:15", "%H:%M")

    time_1 = generate_times(start_shift_1, 8, max_cadence_1)
    time_2 = generate_times(start_shift_2, 8, max_cadence_2)

    # === Write initial Excel to disk ===
    with pd.ExcelWriter(combined_filename, engine='openpyxl') as writer:
        for unit in units_to_plan:
            df_output, df_dock_space, side_lane, lane, rack = run_analysis(uploaded_file, unit, inventory_on_hand, dock_on_hand, move_order_prev_Day, time_1, time_2)
            unit_lanes[unit] = (side_lane, lane, rack)

            df_output.to_excel(writer, sheet_name=f"{unit}-Delivery", index=False)
            df_dock_space.to_excel(writer, sheet_name=f"{unit}-DockSpace", index=False)

    # === Apply highlights ===
    for unit in units_to_plan:
        side_lane, lane, rack = unit_lanes[unit]
        highlight_side_lane(combined_filename, f"{unit}-Delivery", side_lane, lane, rack)
        highlight_side_lane(combined_filename, f"{unit}-DockSpace", side_lane, lane, rack)

    # === Add summary deliveries ===
    summary_df = summary_delivery(combined_filename, units_to_plan)
    with pd.ExcelWriter(combined_filename, mode='a', engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='Summary_Deliveries', index=False)

    # === Read back to memory for download ===
    with open(combined_filename, "rb") as f:
        output_buffer = BytesIO(f.read())

    st.success("Download your plan below.")
    st.download_button(
        label="Download Delivery Plan",
        data=output_buffer,
        file_name=combined_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
