    # Extract delivery columns

from analysis_helpers import build_delivery_plan, build_dock_space_analysis, append_summary_rows
from drive_specifics import get_lane_material, get_drive_unit_buffer_rate

import pandas as pd
import math
from datetime import datetime
from openpyxl import load_workbook
import io


def run_analysis(uploaded_file, input_drive_unit, inventory_on_hand=None, dock_on_hand=None, move_order_prev=None, time_1=None, time_2=None):

    "Main function to run the analysis."

    #Read in the Excel file 
    wb = load_workbook(uploaded_file, data_only=True)
    sheet_name = f'Inbound-{input_drive_unit}'
    ws = wb[sheet_name] 

    #Get side lane and lane material 
    side_lane, lane, rack = get_lane_material(input_drive_unit)
    buffer_rate = get_drive_unit_buffer_rate(input_drive_unit)

    # Read Dock Parameters
    box_dock_space = ws['D2'].value
    pallet_per_lane = ws['D3'].value
    side_lane_pallet = ws['D4'].value

    # Read in Number of Lines
    num_lines_1 = ws['B9'].value
    num_lines_2 = ws['B10'].value

    # Calculate shift hours 
    shift_1_hours = (datetime.strptime("14:15", "%H:%M") - datetime.strptime("6:15", "%H:%M")).seconds / 3600
    shift_2_hours = (datetime.strptime("23:15", "%H:%M") - datetime.strptime("15:15", "%H:%M")).seconds / 3600

    #  Read BOM data
    df_bom = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=15, header=None)
    df_bom.columns = [
        'Part Number', 'Description', 'Quantity / Unit', 'Needed per day',
        'Quantity Needed for Shift 1', 'Quantity Needed for Shift 2',
        'Pallets Utilized for Shift 1', 'Pallets Utilized for Shift 2',
        'Consumption Rate Units/ Hour Shift 1', 'Consumption Rate / Hour Shift 2',
        'Standard Pack Size', 'Package Type', 'Maximum Storage on Lineside',
        'Minimum Storage on Lineside', 'On-hand qty', 'In-Transit QTY', 'On-hand on dock', 'On-hand QTY at Lineside', 'Total Move Order - Prev Day',
        'Total QTY Needed Previous Day'
    ]

    # Generate delivery plan
    df_output = build_delivery_plan(df_bom, inventory_on_hand, move_order_prev, time_1, time_2, shift_1_hours, shift_2_hours, buffer_rate)
    # Build dock space analysis
    df_dock_space = build_dock_space_analysis(df_bom, df_output, dock_on_hand, time_1, time_2, shift_1_hours, shift_2_hours, num_lines_1, num_lines_2)   
    df_dock_space['Part Number'] = df_dock_space['Part Number'].astype(str).str.strip()
    side_lane = [str(x).strip() for x in side_lane]
    lane = [str(x).strip() for x in lane]
    rack = [str(x).strip() for x in rack] 
    df_dock_space = append_summary_rows(df_dock_space, box_dock_space, side_lane_pallet, pallet_per_lane, side_lane, lane, rack)

    return df_output, df_dock_space, side_lane, lane, rack


