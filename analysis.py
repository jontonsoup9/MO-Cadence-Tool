    # Extract delivery columns

from analysis_helpers import get_lane_material, build_delivery_plan, build_dock_space_analysis, append_summary_rows

import pandas as pd
import math
from datetime import datetime
from openpyxl import load_workbook
import io


def run_analysis(uploaded_file, input_drive_unit, inventory_on_hand=None, dock_on_hand=None, time_1=None, time_2=None):

    "Main function to run the analysis."

    #Read in the Excel file 
    wb = load_workbook(uploaded_file, data_only=True)
    sheet_name = f'Inbound-{input_drive_unit}'
    ws = wb[sheet_name] 

    #Get side lane and lane material 
    side_lane, lane = get_lane_material(input_drive_unit)

    # Read Dock Parameters
    box_dock_space = ws['D2'].value
    pallet_per_lane = ws['D3'].value
    side_lane_pallet = ws['D4'].value

    # Calculate shift hours 
    shift_1_hours = (datetime.strptime("14:15", "%H:%M") - datetime.strptime("6:15", "%H:%M")).seconds / 3600
    shift_2_hours = (datetime.strptime("23:15", "%H:%M") - datetime.strptime("15:15", "%H:%M")).seconds / 3600

    #  Read BOM data
    df_bom = pd.read_excel(uploaded_file, sheet_name=sheet_name, skiprows=15, header=None)
    df_bom.columns = [
        'Part Number', 'Description', 'Quantity / Unit', 'Needed per day',
        'Quantity Needed for Shift 1', 'Quantity Needed for Shift 2',
        'Pallets Utilized for Shift 1', 'Pallets Utilized for Shift 2',
        'Consumption Rate Units/ Hour Shift 1', 'Consumption Rate / Hour Shift 2',
        'Standard Pack Size', 'Package Type', 'Maximum Storage on Lineside',
        'Minimum Storage on Lineside', 'On-hand qty', 'QTY vs Shift 1', 'On-hand on dock', 'On-hand QTY at Lineside'
    ]

    # Generate delivery plan
    df_output = build_delivery_plan(df_bom, inventory_on_hand, time_1, time_2, shift_1_hours, shift_2_hours)
    # Build dock space analysis
    df_dock_space = build_dock_space_analysis(df_bom, df_output, dock_on_hand, time_1, time_2, shift_1_hours, shift_2_hours)    
    df_dock_space = append_summary_rows(df_dock_space, box_dock_space, side_lane_pallet, pallet_per_lane, side_lane, lane)

    return df_output, df_dock_space, side_lane, lane


