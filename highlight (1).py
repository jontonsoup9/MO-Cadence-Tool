import pandas as pd
import math
from datetime import datetime
from openpyxl import load_workbook
import io
from openpyxl.styles import PatternFill

def add_color_key_to_side(ws):
    sand_fill = PatternFill(start_color="F4E1C1", end_color="F4E1C1", fill_type="solid")
    sand_blue_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
    sand_green_fill = PatternFill(start_color="C1D1C1", end_color="C1D1C1", fill_type="solid")
    sand_gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    legend_items = [
        ("Side Lane Part / Usage", sand_fill),
        ("Lane Part / Usage", sand_blue_fill),
        ("Totals", sand_green_fill),
        ("Racks", sand_gray_fill)
    ]

    # Find starting column (2 columns after last column of data)
    start_col = ws.max_column + 2
    start_row = 2  # You can change where vertically you want it

    for i, (label, fill) in enumerate(legend_items):
        cell_label = ws.cell(row=start_row + i, column=start_col, value=label)
        cell_color = ws.cell(row=start_row + i, column=start_col + 1)
        cell_color.fill = fill

#   Styling Helper Functions 
def highlight_side_lane(file_name, sheet_name, side_lane, lane, rack):
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define sand fill
    sand_fill = PatternFill(start_color="F4E1C1", end_color="F4E1C1", fill_type="solid")
    sand_blue_fill = PatternFill(start_color="B0C4DE", end_color="B0C4DE", fill_type="solid")
    sand_green_fill = PatternFill(start_color="C1D1C1", end_color="C1D1C1", fill_type="solid")
    sand_gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid") 
    # Get header to find Part Number column index
    header = [cell.value for cell in ws[1]]
    part_num_col_idx = header.index("Part Number") + 1  # openpyxl columns are 1-based
    pkg_type_col_idx = header.index("Package Type") + 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        part_value = str(row[part_num_col_idx - 1].value).strip()
        pkg_type_value = str(row[pkg_type_col_idx - 1].value).strip()

        if part_value in side_lane or "SIDE LANE % Usage" in part_value or "TOTAL - SIDE LANE" in part_value:
            for cell in row:
                cell.fill = sand_fill
        elif part_value in lane or part_value.startswith("LANE % -"):
            # Highlight lane rows sand blue
            for cell in row:
                cell.fill = sand_blue_fill
        elif "TOTAL - BOX" in part_value or "TOTAL - PALLET" in part_value:
            for cell in row: 
                cell.fill  = sand_green_fill
        elif part_value in rack or "PALLET RACK % Usage" in part_value or "TOTAL - RACK BOXES" in part_value:
            for cell in row:
                cell.fill = sand_gray_fill
    
    add_color_key_to_side(ws)
        
    wb.save(file_name) 