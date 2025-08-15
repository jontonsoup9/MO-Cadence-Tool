
from analysis import run_analysis
from highlight import highlight_side_lane
from delivery_helpers import generate_times
from summary import summary_delivery

import pandas as pd
import math
from datetime import datetime
from openpyxl import load_workbook
import numpy as np
import io
import tkinter as tk
from tkinter import filedialog
from openpyxl.styles import PatternFill

if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select BOM Excel File",
        filetypes=[("Excel files", "*.xlsx *.xlsm")]
    )

    wb = load_workbook(file_path, data_only=True)
    print("Available sheets:", wb.sheetnames)

    print("Available Drive Units: Proteus, Hercules, Megasus")
    input_units = input("Enter Drive Units to plan (comma-separated, or 'all' for all units): ").strip().lower()
    valid_units = ['Hercules', 'Megasus', 'Proteus']

    if input_units == 'all':
        units_to_plan = valid_units
    else:
        units_to_plan = [u.strip().capitalize() for u in input_units.split(",")]
        for u in units_to_plan:
            if u not in valid_units:
                raise ValueError(f"Invalid drive unit: {u}")

    first_unit = units_to_plan[0]
    try:
        df_bom_init = pd.read_excel(file_path, sheet_name=f"Inbound-{first_unit}", skiprows=15, header=None)
    except ValueError:
        raise ValueError(f"Sheet 'Inbound-{first_unit}' not found in {file_path}. Check sheet names.")

    df_bom_init.columns = ['Part Number', 'Description', 'Quantity / Unit', 'Needed per day',
        'Quantity Needed for Shift 1', 'Quantity Needed for Shift 2',
        'Pallets Utilized for Shift 1', 'Pallets Utilized for Shift 2',
        'Consumption Rate Units/ Hour Shift 1', 'Consumption Rate / Hour Shift 2',
        'Standard Pack Size', 'Package Type', 'Maximum Storage on Lineside',
        'Minimum Storage on Lineside', 'On-hand qty', 'In-Transit QTY', 'On-hand on dock', 'On-hand QTY at Lineside', 'Total Move Order - Prev Day', 'Total QTY Needed Previous Day']
    
    #Creating dictionary for inventory on hand across all units
    inventory_on_hand = {
        part: qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['On-hand qty'])
    }
    dock_on_hand = { 
        part:qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['On-hand on dock'])
    }
    move_order_prev_Day = {
        part:qty for part, qty in zip(df_bom_init['Part Number'], df_bom_init['Total Move Order - Prev Day'])
    }

    # Determine Max Cadence to align deliveries per unit 
    cadence_shift_1_list = []
    cadence_shift_2_list = []
    for unit in units_to_plan:
        wb = load_workbook(file_path, data_only=True)
        ws = wb[f'Inbound-{unit}']
        c1 = ws['B5'].value
        c2 = ws['B13'].value

        c1 = int(c1) if c1 is not None and str(c1).isdigit() else 0
        c2 = int(c2) if c2 is not None and str(c2).isdigit() else 0

        cadence_shift_1_list.append(c1)
        cadence_shift_2_list.append(c2)

    max_cadence_1 = max(cadence_shift_1_list)
    max_cadence_2 = max(cadence_shift_2_list)

    # Generate aligned delivery slots
    start_shift_1 = datetime.strptime("6:15", "%H:%M")
    end_shift_1 = datetime.strptime("15:00", "%H:%M")
    start_shift_2 = datetime.strptime("15:00", "%H:%M")
    end_shift_2 = datetime.strptime("23:15", "%H:%M")

    time_1 = generate_times(start_shift_1, 8, max_cadence_1)
    time_2 = generate_times(start_shift_2, 8, max_cadence_2)


    combined_filename = "Delivery_Inventory_Plan_All.xlsx"
    unit_lanes = {}

    with pd.ExcelWriter(combined_filename, engine='openpyxl') as writer:
        for unit in units_to_plan:
            print(f"Running analysis for {unit}...")
            df_output, df_dock_space, side_lane, lane, rack = run_analysis(file_path, unit, inventory_on_hand, dock_on_hand, move_order_prev_Day, time_1, time_2)
            unit_lanes[unit] = (side_lane, lane, rack)

            df_output.to_excel(writer, sheet_name=f"{unit}-Delivery", index=False)
            df_dock_space.to_excel(writer, sheet_name=f"{unit}-DockSpace", index=False)

    for unit in units_to_plan:
        side_lane, lane, rack = unit_lanes[unit]
        highlight_side_lane(combined_filename, f"{unit}-Delivery", side_lane, lane, rack)
        highlight_side_lane(combined_filename, f"{unit}-DockSpace", side_lane, lane, rack)

summary_df = summary_delivery(combined_filename, units_to_plan)
with pd.ExcelWriter(combined_filename, mode='a', engine='openpyxl') as writer:
    summary_df.to_excel(writer, sheet_name='Summary_Deliveries', index=False)
